#!/usr/bin/env python3
"""
Standalone script to create an example Excel template with named ranges and FreeMarker placeholders.
Run: python3 create_example_excel.py
Output: config-repo/common-templates/templates/example.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Sheet1"

# Header row
sheet['A1'] = "Customer Name"
sheet['B1'] = "Date"
sheet['A1'].font = Font(bold=True)
sheet['B1'].font = Font(bold=True)

# FreeMarker placeholder in row 2 (data row)
sheet['A2'] = "${customerName}"
sheet['B2'] = "${orderDate}"

# Table header for items (row 4)
sheet['A4'] = "Item"
sheet['B4'] = "Qty"
sheet['C4'] = "Price"
for col in ['A', 'B', 'C']:
    sheet[col + '4'].font = Font(bold=True)
    sheet[col + '4'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Create named ranges using DefinedName objects
from openpyxl.workbook.defined_name import DefinedName

# Single cell: customerName -> A2
wb.defined_names["customerName"] = DefinedName("customerName", "$'Sheet1'!$A$2")

# Single cell: orderDate -> B2
wb.defined_names["orderDate"] = DefinedName("orderDate", "$'Sheet1'!$B$2")

# Table area: items -> A5:C10 (rows for data)
wb.defined_names["items"] = DefinedName("items", "$'Sheet1'!$A$5:$C$10")

# Adjust column widths
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 12

# Create output directory if needed
output_dir = "config-repo/common-templates/templates"
os.makedirs(output_dir, exist_ok=True)

# Save the workbook
output_path = os.path.join(output_dir, "example.xlsx")
wb.save(output_path)

print(f"✓ Created: {output_path}")
print(f"  - Named range 'customerName' -> A2 (FreeMarker: ${{customerName}})")
print(f"  - Named range 'orderDate' -> B2 (FreeMarker: ${{orderDate}})")
print(f"  - Named range 'items' -> A5:C10 (table area)")
print()
print("Next: call the API endpoints:")
print("  1. POST /api/documents/fill-excel")
print("  2. POST /api/documents/verify-excel")
